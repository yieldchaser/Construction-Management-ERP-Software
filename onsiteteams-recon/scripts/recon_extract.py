#!/usr/bin/env python3
import os, json, csv, io, sys
from pathlib import Path
from datetime import datetime
import openpyxl
import pdfplumber

BASE = Path(r"C:\Users\Dell\Github\Construction-Management-ERP-Software\onsiteteams-recon")
REPORT = BASE / "recon-extraction-reports"

def rel(path):
    return str(path.relative_to(BASE))

def analyze_excel(path):
    info = {"path": rel(path), "size_kb": round(path.stat().st_size/1024,2), "type": "xlsx"}
    try:
        data = path.read_bytes()
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheets = []
        for sname in wb.sheetnames:
            ws = wb[sname]
            max_row = ws.max_row
            max_col = ws.max_column
            headers = []
            sample_rows = []
            for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(6, max_row), values_only=False)):
                vals = [str(c.value)[:120] if c.value is not None else "" for c in row]
                if ridx == 0:
                    headers = vals[:min(15, max_col)]
                elif ridx <= 3:
                    sample_rows.append(vals)
            sheets.append({
                "name": sname,
                "max_row": max_row,
                "max_col": max_col,
                "headers": headers,
                "sample_rows": sample_rows
            })
        info["sheets"] = sheets
        wb.close()
    except Exception as e:
        info["error"] = str(e)
    return info

def analyze_csv(path):
    info = {"path": rel(path), "size_kb": round(path.stat().st_size/1024,2), "type": "csv"}
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            content = f.read()
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
        info["rows"] = len(rows)
        info["cols"] = len(rows[0]) if rows else 0
        info["headers"] = rows[0][:15] if rows else []
        info["sample_rows"] = rows[1:5] if len(rows)>1 else []
    except Exception as e:
        info["error"] = str(e)
    return info

def analyze_pdf(path):
    info = {"path": rel(path), "size_kb": round(path.stat().st_size/1024,2), "type": "pdf"}
    try:
        with pdfplumber.open(path) as pdf:
            info["pages"] = len(pdf.pages)
            text_snippets = []
            tables = []
            for i, page in enumerate(pdf.pages[:4]):
                t = page.extract_text()
                if t:
                    text_snippets.append(f"Page {i+1}: {t[:1200]}")
                tbl = page.extract_tables()
                if tbl:
                    for ti, table in enumerate(tbl[:2]):
                        tables.append({
                            "page": i+1,
                            "table_index": ti+1,
                            "rows": len(table),
                            "cols": len(table[0]) if table else 0,
                            "sample": table[:4]
                        })
            info["text_preview"] = text_snippets
            info["tables_preview"] = tables
    except Exception as e:
        info["error"] = str(e)
    return info

def analyze_har(path):
    info = {"path": rel(path), "size_mb": round(path.stat().st_size/1024/1024,2), "type": "har"}
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
        info["raw_length"] = len(text)
        info["preview"] = text[:6000]
    except Exception as e:
        info["error"] = str(e)
    return info

def write_report(filename, title, items):
    lines = [f"# {title}\n", f"Generated: {datetime.now().isoformat()}\n", f"Total items: {len(items)}\n", "---\n"]
    for item in items:
        lines.append(f"## {item['path']}\n")
        size_val = item.get('size_kb', item.get('size_mb', 'N/A'))
        size_unit = 'KB' if 'size_kb' in item else 'MB'
        lines.append(f"- Size: {size_val} {size_unit}")
        lines.append(f"- Type: {item.get('type', 'unknown')}")
        if "error" in item:
            lines.append(f"- **ERROR**: {item['error']}")
            lines.append("\n")
            continue
        if item.get("type") == "xlsx":
            for s in item.get("sheets", []):
                lines.append(f"\n### Sheet: {s['name']} (rows={s['max_row']}, cols={s['max_col']})")
                lines.append(f"- Headers: {s['headers']}")
                for r in s.get("sample_rows", []):
                    lines.append(f"  - {r}")
        elif item.get("type") == "csv":
            lines.append(f"- Rows: {item.get('rows', 'N/A')}, Cols: {item.get('cols', 'N/A')}")
            lines.append(f"- Headers: {item.get('headers', [])}")
            for r in item.get("sample_rows", []):
                lines.append(f"  - {r}")
        elif item.get("type") == "pdf":
            lines.append(f"- Pages: {item.get('pages', 'N/A')}")
            for s in item.get("text_preview", []):
                lines.append(f"\n{s[:1200]}")
            for t in item.get("tables_preview", []):
                lines.append(f"\nTable (page {t['page']}, {t['rows']}x{t['cols']}):")
                for row in t.get("sample", [])[:4]:
                    lines.append(f"  {row}")
        elif item.get("type") == "har":
            lines.append(f"- Raw length: {item.get('raw_length', 'N/A')} chars")
            lines.append(f"\n```\n{item.get('preview', '')[:6000]}\n```")
        lines.append("\n")
    (REPORT / filename).write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {filename} ({len(items)} items)")

def main():
    REPORT.mkdir(exist_ok=True)
    root = BASE.parent
    excel_files = []
    for d in [BASE / "Extra HAR + Image Recon" / "Downloaded Excels Login Site", root]:
        for ext in ["*.xlsx", "*.csv"]:
            excel_files.extend(d.glob(ext))
    seen = set()
    unique_excel = []
    for f in excel_files:
        if f.name not in seen:
            seen.add(f.name)
            unique_excel.append(f)
    unique_excel.sort(key=lambda x: x.name)
    
    excel_items = []
    for f in unique_excel:
        if f.suffix.lower() == ".csv":
            with open(f, "rb") as fp:
                h = fp.read(4)
            if h.startswith(b"PK"):
                excel_items.append(analyze_excel(f))
            else:
                excel_items.append(analyze_csv(f))
        else:
            excel_items.append(analyze_excel(f))
    write_report("01-EXCEL-AND-CSV-REPORT.md", "Excel / CSV Deep Recon", excel_items)
    
    har_files = sorted((BASE / "Extra HAR + Image Recon" / "Har Files").glob("*.har"))
    har_items = [analyze_har(f) for f in har_files]
    write_report("02-HAR-REPORT.md", "HAR Files Deep Recon", har_items)
    
    pdf_files = sorted((BASE / "Extra HAR + Image Recon" / "Main Site PDFs").glob("*.pdf"))
    pdf_items = [analyze_pdf(f) for f in pdf_files]
    write_report("03-PDF-REPORT.md", "PDF Deep Recon (All 45 PDFs)", pdf_items)

if __name__ == "__main__":
    main()
