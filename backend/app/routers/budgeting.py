import io
from uuid import UUID
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Response, status
from sqlalchemy.orm import Session
from sqlalchemy import select
from openpyxl import load_workbook
from app.database import get_db
from app.auth import get_current_user, verify_project_access, get_company_membership
from app.models import BOQItem, BOQDocument, ProjectBudget, Project, Bill, LibraryParty, Task, User
from app.workflow_controls import get_default_terms
from app.utils.pdf_generator import generate_document_pdf
from app.utils.document_pdf import resolve_pdf_branding
from pydantic import BaseModel, Field

router = APIRouter(
    prefix="/budgeting",
    tags=["Budgeting & BOQ"],
    dependencies=[Depends(get_current_user)]
)

class BOQItemResponse(BaseModel):
    id: UUID
    project_id: UUID
    section_name: Optional[str] = None
    item_name: str
    unit: str
    quantity: float
    rate: float
    supply_rate: float
    installation_rate: float
    amount: float
    quantity_float_limit: int

    class Config:
        from_attributes = True

class BudgetAllocationRequest(BaseModel):
    project_id: UUID
    material_budget: float = 0.0
    labour_budget: float = 0.0
    subcon_budget: float = 0.0
    equipment_budget: float = 0.0

class BudgetResponse(BaseModel):
    id: UUID
    project_id: UUID
    material_budget: float
    labour_budget: float
    subcon_budget: float
    equipment_budget: float

    class Config:
        from_attributes = True

@router.get("/boq", response_model=List[BOQItemResponse])
def get_boq_items(project_id: UUID, boq_document_id: Optional[UUID] = None, db: Session = Depends(get_db), _: None = Depends(verify_project_access)):
    # Check if project exists
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    q = db.query(BOQItem).filter(BOQItem.project_id == project_id)
    if boq_document_id is not None:
        q = q.filter(BOQItem.boq_document_id == boq_document_id)
    items = q.all()
    # Cast Numeric types to floats for response model compatibility
    result = []
    for item in items:
        result.append(BOQItemResponse(
            id=item.id,
            project_id=item.project_id,
            section_name=item.section_name,
            item_name=item.item_name,
            unit=item.unit,
            quantity=float(item.quantity),
            rate=float(item.rate),
            supply_rate=float(item.supply_rate),
            installation_rate=float(item.installation_rate),
            amount=float(item.amount or 0.0) if item.amount is not None else float(item.quantity) * (float(item.rate) + float(item.supply_rate) + float(item.installation_rate)),
            quantity_float_limit=item.quantity_float_limit
        ))
    return result

@router.post("/boq/import", status_code=status.HTTP_201_CREATED)
async def import_boq(
    project_id: UUID = Form(...),
    file: UploadFile = File(...),
    boq_document_id: Optional[UUID] = Form(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    # project_id here comes from multipart form data (not the URL path/query),
    # so it can't share a value with a plain Depends(verify_project_access)
    # sub-dependency; verify membership inline once the project is loaded instead.
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    get_company_membership(db, current_user, project.company_id)

    if not file.filename.endswith(('.xlsx', '.xlsm')):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xlsm Excel files are supported")

    try:
        contents = await file.read()
        wb = load_workbook(filename=io.BytesIO(contents), read_only=True)
        sheet = wb.active
        if not sheet:
            raise HTTPException(status_code=400, detail="Excel file is empty")

        # Parse header row (assume first row has headers)
        headers = {}
        header_row = next(sheet.iter_rows(values_only=True))
        for idx, val in enumerate(header_row):
            if val:
                headers[str(val).strip().lower()] = idx

        # Required columns check
        name_col = next((headers[k] for k in ["item_name", "item", "description", "name"] if k in headers), None)
        qty_col = next((headers[k] for k in ["qty", "quantity"] if k in headers), None)
        unit_col = headers.get("unit")
        rate_col = next((headers[k] for k in ["rate", "unit_rate"] if k in headers), None)
        
        # Optional split rates
        supply_rate_col = headers.get("supply_rate")
        install_rate_col = next((headers[k] for k in ["installation_rate", "install_rate"] if k in headers), None)
        section_col = next((headers[k] for k in ["section", "section_name"] if k in headers), None)

        if name_col is None or qty_col is None or unit_col is None or (rate_col is None and supply_rate_col is None):
            raise HTTPException(
                status_code=400,
                detail="Invalid headers. Excel must contain: Description/Item Name, Qty, Unit, and Rate (or Supply Rate)."
            )

        imported_count = 0
        total_amount = 0.0

        # Iterate rows
        is_first = True
        for row in sheet.iter_rows(values_only=True):
            if is_first:
                is_first = False
                continue

            item_name = row[name_col]
            if not item_name:
                continue # Skip blank lines

            qty_val = row[qty_col]
            unit_val = row[unit_col]
            rate_val = row[rate_col] if rate_col is not None else 0.0
            supply_val = row[supply_rate_col] if supply_rate_col is not None else 0.0
            install_val = row[install_rate_col] if install_rate_col is not None else 0.0

            try:
                quantity = float(qty_val) if qty_val is not None else 0.0
                rate = float(rate_val) if rate_val is not None else 0.0
                supply_rate = float(supply_val) if supply_val is not None else 0.0
                installation_rate = float(install_val) if install_val is not None else 0.0
            except ValueError:
                continue # Skip rows with non-numeric qty/rates

            section_name = str(row[section_col]).strip() if (section_col is not None and row[section_col]) else None
            unit = str(unit_val).strip() if unit_val else "Nos"

            # Enforce quantity float limit (rounding base)
            float_limit = 2
            if unit.lower() in ("kg", "ton", "t", "steel"):
                float_limit = 3
            elif unit.lower() in ("no", "nos", "brick", "bag", "bags"):
                float_limit = 0

            quantity = round(quantity, float_limit)
            amount = quantity * (rate + supply_rate + installation_rate)
            total_amount += amount

            boq_item = BOQItem(
                project_id=project_id,
                boq_document_id=boq_document_id,
                section_name=section_name,
                item_name=str(item_name).strip(),
                unit=unit,
                quantity=quantity,
                rate=rate,
                supply_rate=supply_rate,
                installation_rate=installation_rate,
                quantity_float_limit=float_limit,
                amount=amount
            )
            db.add(boq_item)
            imported_count += 1

        db.commit()
        return {
            "success": True,
            "imported_count": imported_count,
            "total_estimated_cost": total_amount
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")

@router.post("/allocation", response_model=BudgetResponse)
def allocate_project_budgets(
    request: BudgetAllocationRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    project = db.query(Project).filter(Project.id == request.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    get_company_membership(db, current_user, project.company_id)

    budget = db.query(ProjectBudget).filter(ProjectBudget.project_id == request.project_id).first()
    if not budget:
        budget = ProjectBudget(project_id=request.project_id)
        db.add(budget)

    budget.material_budget = request.material_budget
    budget.labour_budget = request.labour_budget
    budget.subcon_budget = request.subcon_budget
    budget.equipment_budget = request.equipment_budget

    db.commit()
    db.refresh(budget)
    return budget


# ─── BOQ Documents (per-client BOQ layer) ────────────────────────────────────
# Invoices raised to the client whose value counts toward "Billed Value".
BILLING_TYPES = {"sales_invoice", "material_sales"}


def _item_amount(i: BOQItem) -> float:
    if i.amount is not None:
        return float(i.amount)
    return float(i.quantity) * (float(i.rate) + float(i.supply_rate) + float(i.installation_rate))


class BOQDocumentCreate(BaseModel):
    project_id: UUID
    title: str
    client_party_id: Optional[UUID] = None
    milestone_done: int = 0
    milestone_total: int = 0
    terms: Optional[str] = None  # Terms & Conditions; defaults to company BOQ Terms on create


class BOQDocumentPatch(BaseModel):
    title: Optional[str] = None
    client_party_id: Optional[UUID] = None
    milestone_done: Optional[int] = None
    milestone_total: Optional[int] = None
    terms: Optional[str] = None


class BOQDocumentResponse(BaseModel):
    id: UUID
    project_id: UUID
    client_party_id: Optional[UUID] = None
    client_name: Optional[str] = None
    title: str
    milestone_done: int
    milestone_total: int
    boq_value: float
    billed_value: float
    physical_progress: float  # 0-100, value-weighted linked-task completion
    item_count: int
    terms: Optional[str] = None

    class Config:
        from_attributes = True


def _build_doc_response(db: Session, d: BOQDocument) -> BOQDocumentResponse:
    items = db.query(BOQItem).filter(BOQItem.boq_document_id == d.id).all()
    boq_value = sum(_item_amount(i) for i in items)

    bills = (
        db.query(Bill)
        .filter(Bill.project_id == d.project_id, Bill.boq_document_id == d.id, Bill.invoice_type.in_(BILLING_TYPES))
        .all()
    )
    billed_value = sum(float(b.total_payable) for b in bills)

    # Physical progress = value-weighted average of linked-task completion.
    item_ids = [i.id for i in items]
    progress_by_item: dict = {}
    if item_ids:
        tasks = db.query(Task).filter(Task.boq_item_id.in_(item_ids)).all()
        for t in tasks:
            progress_by_item.setdefault(t.boq_item_id, []).append(float(t.progress or 0))
    num = 0.0
    den = 0.0
    for i in items:
        amt = _item_amount(i)
        ps = progress_by_item.get(i.id)
        pct = (sum(ps) / len(ps)) if ps else 0.0
        num += amt * pct
        den += amt
    physical_progress = round(num / den, 2) if den > 0 else 0.0

    party = db.query(LibraryParty).filter(LibraryParty.id == d.client_party_id).first() if d.client_party_id else None
    return BOQDocumentResponse(
        id=d.id,
        project_id=d.project_id,
        client_party_id=d.client_party_id,
        client_name=party.name if party else None,
        title=d.title,
        milestone_done=d.milestone_done,
        milestone_total=d.milestone_total,
        boq_value=round(boq_value, 2),
        billed_value=round(billed_value, 2),
        physical_progress=physical_progress,
        item_count=len(items),
        terms=d.terms,
    )


@router.get("/boq-documents", response_model=List[BOQDocumentResponse])
def list_boq_documents(project_id: UUID, db: Session = Depends(get_db), _: None = Depends(verify_project_access)):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    docs = db.query(BOQDocument).filter(BOQDocument.project_id == project_id).all()
    return [_build_doc_response(db, d) for d in docs]


@router.post("/boq-documents", response_model=BOQDocumentResponse, status_code=201)
def create_boq_document(req: BOQDocumentCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    project = db.query(Project).filter(Project.id == req.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    get_company_membership(db, current_user, project.company_id)
    if req.client_party_id:
        party = db.query(LibraryParty).filter(LibraryParty.id == req.client_party_id).first()
        if not party:
            raise HTTPException(status_code=404, detail="Client party not found")
    doc = BOQDocument(
        project_id=req.project_id,
        client_party_id=req.client_party_id,
        title=req.title,
        milestone_done=req.milestone_done,
        milestone_total=req.milestone_total,
        # Settings -> Terms & Conditions -> BOQ Terms: pre-fill the company
        # default when the caller doesn't supply their own terms.
        terms=req.terms
        if req.terms
        else get_default_terms(db, project.company_id, "boq"),
    )
    db.add(doc)
    db.commit()
    db.refresh(doc)
    return _build_doc_response(db, doc)


@router.patch("/boq-documents/{doc_id}", response_model=BOQDocumentResponse)
def patch_boq_document(doc_id: UUID, req: BOQDocumentPatch, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    doc = db.query(BOQDocument).filter(BOQDocument.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="BOQ document not found")
    project = db.query(Project).filter(Project.id == doc.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    get_company_membership(db, current_user, project.company_id)
    if req.title is not None:
        doc.title = req.title
    if req.client_party_id is not None:
        doc.client_party_id = req.client_party_id
    if req.milestone_done is not None:
        doc.milestone_done = req.milestone_done
    if req.milestone_total is not None:
        doc.milestone_total = req.milestone_total
    if req.terms is not None:
        doc.terms = req.terms
    db.commit()
    db.refresh(doc)
    return _build_doc_response(db, doc)


@router.get("/boq-documents/{doc_id}/pdf")
def get_boq_document_pdf(doc_id: UUID, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    doc = db.query(BOQDocument).filter(BOQDocument.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="BOQ document not found")
    # Tenant check: the document's project belongs to a company the caller is a member of.
    project = db.query(Project).filter(Project.id == doc.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    get_company_membership(db, current_user, project.company_id)

    company_name, custom_banner = resolve_pdf_branding(db, project.company_id, project)
    party = db.query(LibraryParty).filter(LibraryParty.id == doc.client_party_id).first() if doc.client_party_id else None
    client_name = party.name if party else "N/A"

    party_lines = [
        f"Client: {client_name}",
        f"Title: {doc.title}",
        f"Milestones: {doc.milestone_done} / {doc.milestone_total}",
    ]

    items = db.query(BOQItem).filter(BOQItem.boq_document_id == doc.id).all()
    table_headers = ["Section", "Item", "Unit", "Qty", "Rate", "Amount"]
    col_widths = [14, 30, 8, 10, 14, 14]
    table_rows = []
    boq_value = 0.0
    for it in items:
        amt = _item_amount(it)
        boq_value += amt
        table_rows.append([
            it.section_name or "",
            it.item_name,
            it.unit,
            str(it.quantity),
            str(it.rate),
            f"{amt:.2f}",
        ])
    if not table_rows:
        table_rows.append(["", "(No line items)", "", "", "", ""])

    totals_lines = [
        f"BOQ Value: {boq_value:.2f}",
        f"Line Items: {len(items)}",
    ]

    pdf_bytes = generate_document_pdf(
        title="Bill of Quantities (BOQ)",
        party_lines=party_lines,
        table_headers=table_headers,
        table_rows=table_rows,
        col_widths=col_widths,
        totals_lines=totals_lines,
        terms=doc.terms,
        company_name=company_name,
        custom_banner=custom_banner,
    )
    filename = f"boq-{doc.title or doc.id}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )
